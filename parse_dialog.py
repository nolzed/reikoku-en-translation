import struct, argparse, json

from unpack_spirit import align_4
from font_mapper import FontMapper

def parse_dialog_text(text_bytes, font_map: FontMapper):
    i = 0
    output = []

    def read_next():
        nonlocal i
        if i < len(text_bytes):
            val = text_bytes[i]
            i += 1
            return val
        return None
    
    def is_next():
        nonlocal i
        if i < len(text_bytes):
            return text_bytes[i]
        return None
    
    while i < len(text_bytes):
        byte = read_next()
        if byte is None:
            break
            
        # Control characters
        # 0x00 - [END] end of string
        # 0x01:0x0a - font code
        # 0x0b - [WAIT_1] wait for any button to be pressed and continue with a new line
        # 0x0c - [FUNC_ID:XX] call a function by index in a table
        # 0x0d - \n
        # 0x0e - [INDENT] sets the indentation for the following lines based on the current position in the line
        #
        # 0x0f - special control
        #        0x00 - [DELAY:XX] frames count
        #        0x01 - [WAIT_2] wait for input without new line
        #        0x02 - [CLEAR] clear window and continue output
        #        0x04 - [FUNC_ADR:0x{addr:04X}] call a function by address
        #
        # 0x20 - " " or [SP:X] if there are more than 2
        # 0x25 - "%" string formatting
        # >= 0x20  - ascii characters that are encoded through "ascii-table.bin"
        
        if byte == 0x00:
            output.append("[END]")
            remaining = text_bytes[i:]
            if remaining:
                hex_str = ''.join(f'{b:02X}' for b in remaining)
                output.append(f"[RAW:{hex_str}]")
            break

        elif 0x01 <= byte <= 0x0A and is_next() is not None:
            high = byte
            low = read_next()
            code = (high << 8) | (low if low is not None else 0)
            output.append(font_map.get_char(code))

        elif byte == 0x0B:
            output.append("[WAIT_1]")

        elif byte == 0x0C and is_next() is not None:
            func_id = read_next()
            output.append(f"[FUNC_ID:{func_id}]")

        elif byte == 0x0D:
            output.append("\n")

        elif byte == 0x0E:
            output.append("[INDENT]")

        elif byte == 0x0F and is_next() is not None:
            subcmd = read_next()
            if subcmd == 0x00:
                delay = read_next()
                output.append(f"[DELAY:{delay}]")
            elif subcmd == 0x01:
                output.append("[WAIT_2]")
                i -= 2
            elif subcmd == 0x02:
                output.append("[CLEAR]")
            elif subcmd == 0x04:
                arg1 = read_next()
                arg2 = read_next()
                if arg1 is not None and arg2 is not None:
                    addr = (arg2 << 8) | arg1
                    output.append(f"[FUNC_ADR:0x{addr:04X}]")
                else:
                    output.append("[FUNC_ADR:??]")
            else:
                output.append(f"[PAUSE:UNKNOWN:{subcmd}]")

        #elif byte == 0x25: # % args
        #    print(text_bytes, output)
        #    output.append(get_ascii_char(byte, TABLE))
        #    output.append(get_ascii_char(read_next(), TABLE))
        
        elif byte == 0x20:
            count = 1
            while is_next() == 0x20:
                read_next()
                count += 1
            if count == 1:
                char = font_map.get_ascii_char(byte)
                output.append(char)
            else:
                output.append(f"[SP:{count}]")
                
        elif 0x20 <= byte:
            char = font_map.get_ascii_char(byte)
            output.append(char if char else f"[UNK1:{byte:02X}]")
        else:
            output.append(f"[UNK2:{byte:02X}]")
            
    return ''.join(output)

def parse_scenario(data, font_map: FontMapper):
    offset = 0
    
    data_1 = struct.unpack_from('<I', data, offset)[0]
    offset += 4
    
    data_2 = struct.unpack_from('<I', data, offset)[0]
    offset += 4
    
    entries_count_1 = struct.unpack_from('<I', data, offset)[0]
    entries_count_1 += entries_count_1%2
    offset += 4
    print(f"Block 1 entries: {entries_count_1} | {hex(offset)}")
    
    block_3_size = struct.unpack_from('<I', data, offset)[0]
    offset += 4
    print(f"Block 3 size: {block_3_size} | {hex(offset)}")

    entries_1 = []
    for i in range(entries_count_1):
        entry = struct.unpack_from('<H', data, offset + i*2)[0]
        entries_1.append(entry)
        
    offset += entries_count_1 * 2
    
    table_size = struct.unpack_from('<I', data, offset)[0]
    table_offsets_count = (table_size-4)//4
    table_offsets = []
    for i in range(table_offsets_count):
        table_offset = struct.unpack_from('<I', data, offset + 4 + i*4)[0]
        table_offsets.append(table_offset)
    
    print(f"Block 3 table_size: {table_size} | {hex(offset)}")
    
    table_entries = []
    table_raw_entries = []
    
    for i in range(len(table_offsets)):
        if i+1 < len(table_offsets):
            size = table_offsets[i+1] - table_offsets[i]
        else:
            size = block_3_size - table_offsets[i]
        entry_offset = offset + table_offsets[i]
        table_entry = data[entry_offset:entry_offset+size]
        table_raw_entries.append(table_entry.hex())
        if struct.unpack_from("<H", table_entry, 0)[0] * 2 + 2 == len(table_entry):
            table_entries.append(f"[RAW:{table_entry.hex()}]")
        else:
            table_entries.append(parse_dialog_text(table_entry, font_map))
        
    offset += block_3_size
    
    print(f"Block 3 table_entries: {len(table_entries)} | {hex(offset)}")
    
    return {
        "data_1": data_1,
        "data_2": data_2,
        "block_1": {
            "count": entries_count_1,
            "entries": entries_1
        },
        "block_3": {
            "size" : block_3_size,
            "table_size" : table_size,
            "table_offsets": table_offsets,
            "table_entries": table_entries,
            "raw_table_entries": table_raw_entries
        },
        "add_block": data[offset:].hex()
    }
    

def parse_dialog(data, font_map: FontMapper):
    offset = 0
    
    main_size = struct.unpack_from('<I', data, offset)[0]
    offset += 4
    
    entries_count_1 = struct.unpack_from('<I', data, offset)[0]
    entries_count_1 += entries_count_1%2
    offset += 4
    entries_1 = []
    for i in range(entries_count_1):
        entry = struct.unpack_from('<H', data, offset + i*2)[0]
        entries_1.append(entry)
        
    offset += entries_count_1 * 2
    print(f"Block 1 entries: {entries_count_1} | {hex(offset)}")
    
    entries_count_2 = struct.unpack_from('<I', data, offset)[0] 
    entries_count_2 += entries_count_2%2
    offset += 4
    entries_2 = []
    print(entries_count_2)
    for i in range(entries_count_2):
        entry = struct.unpack_from('<H', data, offset + i*2)[0]
        entries_2.append(entry)
        
    offset += entries_count_2 * 2
    print(f"Block 2 entries: {entries_count_2} | {hex(offset)}")
    
    block_3_size = struct.unpack_from('<I', data, offset)[0]
    offset += 4
    
    print(f"Block 3 size: {block_3_size} | {hex(offset)}")
    
    table_size = struct.unpack_from('<I', data, offset)[0]
    
    table_offsets_count = (table_size-4)//4
    table_offsets = []
    for i in range(table_offsets_count):
        table_offset = struct.unpack_from('<I', data, offset + 4 + i*4)[0]
        table_offsets.append(table_offset)
    
    print(f"Block 3 table_size: {table_size} | {hex(offset)}")
    
    table_entries = []
    table_raw_entries = []
    
    for i in range(len(table_offsets)):
        if i+1 < len(table_offsets):
            size = table_offsets[i+1] - table_offsets[i]
        else:
            size = block_3_size - table_offsets[i] - 4
        entry_offset = offset + table_offsets[i]
        table_entry = data[entry_offset:entry_offset+size]
        table_raw_entries.append(table_entry.hex())
        if struct.unpack_from("<H", table_entry, 0)[0] * 2 + 2 == len(table_entry):
            table_entries.append(f"[RAW:{table_entry.hex()}]")
        else:
            table_entries.append(parse_dialog_text(table_entry, font_map))
        
    offset += block_3_size
    
    print(f"Block 3 table_entries: {len(table_entries)} | {hex(offset)}")
    
    return {
        "main_size": main_size,
        "block_1": {
            "count": entries_count_1,
            "entries": entries_1
        },
        "block_2": {
            "count": entries_count_2,
            "entries": entries_2
        },
        "block_3": {
            "size" : block_3_size,
            "table_size" : table_size,
            "table_offsets": table_offsets,
            "table_entries": table_entries,
            "raw_table_entries": table_raw_entries
        },
        "add_block": data[offset:].hex()
    }

import csv
def export_to_csv(entries, filename):
    with open(filename, mode='w', newline='', encoding='shift_jis', errors='replace') as file:
        writer = csv.writer(file, quoting=csv.QUOTE_ALL)
        for item in entries:
            writer.writerow([item])

from openpyxl import Workbook
from openpyxl.styles import Alignment

def export_to_excel(entries, filename, font_line_height=15):
    wb = Workbook()
    ws = wb.active

    for idx, item in enumerate(entries, start=1):
        cell = ws.cell(row=idx, column=1, value=item)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        # count the number of lines by \n
        lines = item.count('\n') + 1

        # Set the row height
        ws.row_dimensions[idx].height = lines * font_line_height

    wb.save(filename)

def export_to_excel_escape(entries, filename, default_row_height=15):
    wb = Workbook()
    ws = wb.active

    for idx, item in enumerate(entries, start=1):
        # Escape line breaks \n -> \\n
        escaped_text = item.replace('\n', '\\n')

        cell = ws.cell(row=idx, column=1, value=escaped_text)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Set the standard row height
        ws.row_dimensions[idx].height = default_row_height

    wb.save(filename)
    
def main():
    parser = argparse.ArgumentParser(description="Parse script file")

    parser.add_argument("file", help="Input dialog file")
    parser.add_argument("dialog_data_out", help="Path to write parsed json data")
    parser.add_argument("excel_out", help="Path to write excel entries")
    
    parser.add_argument(
        "--font_table",
        default="./font/font-table.txt",
        help="Path to font-table.txt"
    )
    parser.add_argument(
        "--ascii_table",
        default="./font/ascii-table.bin",
        help="Path to ascii-table.bin"
    )
    args = parser.parse_args()
    
    with open(args.file, 'rb') as f1:
        data = f1.read()

    dialog_data = parse_dialog(data, FontMapper(args.ascii_table, args.font_table)) # parse_scenario
    
    with open(args.dialog_data_out, 'w', encoding='utf-8') as out:
        json.dump(dialog_data, out, indent=2, ensure_ascii=False)
    
    print("[+] Json dialog_data extracted to:", args.dialog_data_out)

    export_to_excel_escape(dialog_data["block_3"]["table_entries"], args.excel_out)   
    print("[+] Excel dialog_data entries extracted to:", args.excel_out)
        
if __name__ == '__main__':
    main()
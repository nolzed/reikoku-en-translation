import struct
import argparse
import json
from openpyxl import load_workbook

from font_mapper import FontMapper
from unpack_spirit import align_4

from pack_script import normalize_text

def pack_database_text(text, font_map: FontMapper):
    """
    Packs a string with control tokens back into a sequence of bytes.
    """
    
    i = 0
    output = bytearray()

    def emit_byte(b):
        output.append(b & 0xFF)

    while i < len(text):
        if text[i] == '[':
            # parses the token to the nearest ]
            end = text.find(']', i)
            if end == -1:
                raise ValueError(f"Unclosed token at pos {i}")
            token = text[i+1:end]
            i = end + 1  # go for ]
            
            # process each token
            if token.startswith('SP:'):
                # repeat space byte 0x20 count times
                count = int(token.split(':', 1)[1])
                for _ in range(count):
                    emit_byte(0x20)
            elif token == 'END':
                emit_byte(0x00)
            elif token == '':
                continue
            
            elif token.startswith('CLUT:'):
                clut = int(token.split(':', 1)[1])
                emit_byte(0x0C)
                emit_byte(0x0C)
                emit_byte(clut + 0x30)
            
            elif token.startswith('IMG:'):
                img_params = token.split(':', 1)[1]
                emit_byte(0x0C)
                emit_byte(0x10)
                for param in img_params:
                    ch = normalize_text(param)
                    emit_byte(font_map.get_ascii_code(ch))

            elif token.startswith('KEYWORD:'):
                keyw_params = token.split(':', 1)[1]
                emit_byte(0x0C)
                emit_byte(0x14)
                for param in keyw_params:
                    ch = normalize_text(param)
                    emit_byte(font_map.get_ascii_code(ch))

            elif token.startswith('UNK3:'):
                code = int(token.split(':',1)[1], 16)
                emit_byte(0x0C)
                emit_byte(code)

            elif token == 'END_PAGE':
                emit_byte(0x0C)
                emit_byte(0x19)
            elif token == 'LINE':
                emit_byte(0x0C)
                emit_byte(0x1E)

            elif token.startswith('UNK'):
                code = int(token.split(':',1)[1], 16)
                emit_byte(code)

            else:
                raise ValueError(f"Unknown token [{token}]")
            
        elif text[i] == '\n':
            emit_byte(0x0D)
            i += 1
        elif text[i] == '゛':
            emit_byte(0xDE)
            i += 1
        elif text[i] == '゜':
            emit_byte(0xDF)
            i += 1
        else:

            ch = normalize_text(text[i])
            i += 1
            
            if font_map.get_ascii_code(ch):
                emit_byte(font_map.get_ascii_code(ch))
                
            # try it in a two-byte table (0x01–0x0A)
            elif font_map.get_code(ch):
                code = font_map.get_code(ch)
                high = (code >> 8) & 0xFF
                low = code & 0xFF
                # stacked as in parse: 0x01–0x0A => high byte, low byte
                emit_byte(high)
                emit_byte(low)
            else:
                raise ValueError(f"Character '{ch}' cannot be encoded. {text}")
                
    return bytes(output)

def build_database(data, font_map: FontMapper):
    out = bytearray()
    offsets = []

    out += b'\x00' * 4 * 5
    current_offset = len(out)
    offsets.append(len(out))

    # 1. Entries Info Table
    info_entries = data['entries_table']
    entries_info_bytes = bytearray()
    for entry in info_entries:
        entries_info_bytes += struct.pack('<HHH', entry['id'], entry['type'], entry['reserved'])
        entries_info_bytes += struct.pack('<HHHHHHHH', *entry['topics'])
    
    out += entries_info_bytes
    current_offset += len(out)
    offsets.append(len(out))

    # 2. Topics Table
    topics_entries = data['topics_table']['entries']
    n_pages = len(topics_entries)

    # Step 1: Placeholder for table size and offset table
    entry_pages_offset_start = len(out)
    pages_offset_table_size = 4 + 4 * n_pages
    out += b'\x00' * pages_offset_table_size

    # Step 2: Sort entries by sorted_index for writing payload in file order
    sorted_entries = sorted(topics_entries, key=lambda x: x['sorted_index'])

    payload = bytearray()
    page_offsets_by_sorted_index = []
    current_offset = entry_pages_offset_start + pages_offset_table_size

    for entry in sorted_entries:
        search_key = pack_database_text(entry['search_key'], font_map)
        title = pack_database_text(entry['title'], font_map)
        pages = pack_database_text(entry['pages'], font_map)

        packed = search_key + title + pages
        page_offsets_by_sorted_index.append((entry['sorted_index'], current_offset))
        payload += packed
        current_offset += len(packed)

    # Step 3: Map sorted_index -> file offset
    sorted_index_to_offset = dict(page_offsets_by_sorted_index)

    # Step 4: Write offsets to table in original offset_table_index order
    # So that offset table matches the original layout
    for i, entry in enumerate(topics_entries):
        sorted_idx = entry['sorted_index']
        entry_offset = sorted_index_to_offset[sorted_idx]
        struct.pack_into('<I', out, entry_pages_offset_start + 4 + i * 4, entry_offset)

    # Step 5: Write actual data
    out += payload
    out += b'\00' * (align_4(len(out)) - len(out))

    # 3. Keyword Table
    keyword_entries = data['keywords_table']['entries']
    offsets.append(len(out))

    kw_offset_start = len(out)
    n_kw = len(keyword_entries)
    kw_table_size = 4 * n_kw
    out += b'\x00' * kw_table_size

    payload = bytearray()
    kw_offsets = []
    current_offset = kw_offset_start + kw_table_size
    for entry in keyword_entries:
        search_key = pack_database_text(entry['search_key'], font_map)
        text = pack_database_text(entry['text'], font_map)
        packed = search_key + text
        kw_offsets.append(current_offset)
        payload += packed
        current_offset += len(packed)

    for i, offset_val in enumerate(kw_offsets):
        struct.pack_into('<I', out, kw_offset_start + i * 4, offset_val)
    out += payload
    out += b'\00' * (align_4(len(out)) - len(out))

    # 4. Search Table
    search_entries = data['search_table']['entries']
    offsets.append(len(out))

    search_offset_start = len(out)
    n_search = len(search_entries)
    search_table_size = 4 * n_search
    out += b'\x00' * search_table_size

    payload = bytearray()
    search_offsets = []
    current_offset = search_offset_start + search_table_size
    for entry in search_entries:
        header = struct.pack('<BB', entry['key_offset'], entry['count'])
        keys = pack_database_text(entry['keys'], font_map)
        packed = header + keys
        search_offsets.append(current_offset)
        payload += packed
        current_offset += len(packed)

    for i, offset_val in enumerate(search_offsets):
        struct.pack_into('<I', out, search_offset_start + i * 4, offset_val)
    out += payload

    # 5. End raw data
    offsets.append(len(out))
    end_raw = bytes.fromhex(data.get('end_raw_data', ''))
    out += end_raw

    print(offsets)
    # Finally, prepend offset table at the beginning
    final_out = struct.pack('<IIIII', *offsets) + out[0x14:]
    return final_out

def import_from_excel_unescape(filename):
    wb = load_workbook(filename)
    ws = wb.active

    entries = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        cell_value = row[0]
        if cell_value is not None:
            # Unescape \\n -> \n
            unescaped_text = cell_value.replace('\\n', '\n')
            entries.append(unescaped_text)

    return entries

def generate_keyword_table(data):
    keywords = {}
    for entry in data["keywords_table"]["entries"]:
        if entry['definition_id']:
            X = entry['id']
            Z = entry['definition_id']
            
            key = f"({X},0,{Z})"
            keywords[key] = entry['text'].replace("[END]", "")
    return keywords

def main():
    parser = argparse.ArgumentParser(description="Pack JSON script into game format")
    parser.add_argument("input_json", help="Input JSON file with script structure (dialog/scenario)")
    parser.add_argument("out_file", help="Output script file")
    parser.add_argument("--excel", help="Path to excel entries sheets")
    parser.add_argument("--font_table", default="./font/font-table.txt", help="Path to font-table.txt")
    parser.add_argument("--ascii_table", default="./font/ascii-table.bin", help="Path to ascii-table.bin")
    args = parser.parse_args()

    # Load JSON
    with open(args.input_json, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Load and insert excel entries
    #if args.excel:
    #    data['block_3']['entries'] = import_from_excel_unescape(args.excel)
    
    # Build binary
    bin_data = build_database(data, FontMapper(args.ascii_table, args.font_table))

    # Write
    with open(args.out_file, 'wb') as f:
        f.write(bin_data)

    print(f"[+] Script file written to: {args.out_file}")

    with open("./database/keywords_table.json", 'w', encoding='utf-8') as out:
        json.dump(generate_keyword_table(data), out, indent=2, ensure_ascii=False)
    print("[+] JSON Keywords Table saved to:", "./database/keywords_table.json")


if __name__ == '__main__':
    main()

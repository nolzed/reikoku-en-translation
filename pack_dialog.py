import struct
import argparse
import json

from font_mapper import FontMapper
from unpack_spirit import align_4

def pack_dialog_text(text, font_map: FontMapper):
    """
    Packs a string with control tokens back into a sequence of bytes.
    """
    
    i = 0
    output = bytearray()

    def emit_byte(b):
        output.append(b & 0xFF)

    while i < len(text):
        if text[i] == '[':
            # parses the token to the nearest ]
            end = text.find(']', i)
            if end == -1:
                raise ValueError(f"Unclosed token at pos {i}")
            token = text[i+1:end]
            i = end + 1  # go for ]
            
            # process each token
            if token.startswith('RAW:'):
                enters = token.split(":")[1]
                output.extend(bytes.fromhex(enters))
            elif token.startswith('SP:'):
                # repeat space byte 0x20 count times
                count = int(token.split(':', 1)[1])
                for _ in range(count):
                    emit_byte(0x20)
            elif token == 'END':
                emit_byte(0x00)
                #break
            elif token == 'WAIT_1':
                emit_byte(0x0B)
            elif token.startswith('FUNC_ID:'):
                func_id = int(token.split(':',1)[1])
                emit_byte(0x0C)
                emit_byte(func_id)
            elif token == '':
                continue
            elif token == 'INDENT':
                emit_byte(0x0E)
            elif token.startswith('DELAY:'):
                delay = int(token.split(':',1)[1])
                emit_byte(0x0F)
                emit_byte(0x00)
                emit_byte(delay)
            elif token == 'WAIT_2':
                emit_byte(0x0F)
                emit_byte(0x01)
            elif token == 'CLEAR':
                emit_byte(0x0F)
                emit_byte(0x02)
            elif token.startswith('FUNC_ADR:'):
                addr_str = token.split(':',1)[1]
                addr = int(addr_str, 16)
                lo = addr & 0xFF
                hi = (addr >> 8) & 0xFF
                emit_byte(0x0F)
                emit_byte(0x04)
                emit_byte(lo)
                emit_byte(hi)
            elif token.startswith('UNK:'):
                code = int(token.split(':',1)[1], 16)
                emit_byte(code)
            else:
                parts = token.split(':')
                if parts[0] == 'PAUSE' and parts[1] == 'UNKNOWN':
                    emit_byte(0x0F)
                    if parts[2] != 'None':
                        sub = int(parts[2])
                        emit_byte(sub)
                else:
                    raise ValueError(f"Unknown token [{token}]")
        elif text[i] == '\n':
            emit_byte(0x0D)
            i += 1
        else:

            ch = normalize_text(text[i])
            i += 1
            
            if font_map.get_ascii_code(ch):
                emit_byte(font_map.get_ascii_code(ch))
                
            # try it in a two-byte table (0x01–0x0A)
            elif font_map.get_code(ch):
                code = font_map.get_code(ch)
                high = (code >> 8) & 0xFF
                low = code & 0xFF
                # stacked as in parse: 0x01–0x0A => high byte, low byte
                emit_byte(high)
                emit_byte(low)
            else:
                raise ValueError(f"Character '{ch}' cannot be encoded. {text}")
                
    return bytes(output)


def build_dialog(data, font_map: FontMapper):
    out = bytearray()
    
    # Block 1: count entries, drop trailing 0 if present
    entries1 = data['block_1']['entries']
    count1 = len(entries1) - 1 if entries1 and entries1[-1] == 0 else len(entries1)
    
    block1_align = b''
    if (count1 * 2) % 4 != 0:
        block1_align = b'\00' * ((count1 * 2) % 4)
     
    # Block 2: count entries, drop trailing 0 if present
    entries2 = data['block_2']['entries']
    count2 = len(entries2) - 1 if entries2 and entries2[-1] == 0 else len(entries2)

    # Block 3: compute offsets and sizes
    table_entries = data['block_3']['table_entries']
    n = len(table_entries)
    
    # offsets start after table_size (4 bytes) + n*4 bytes
    base_offset = 4 + 4 * n
    offsets = []
    current = base_offset
    payload_bytes = []
    for entry in table_entries:
        offsets.append(current)
        packed = pack_dialog_text(entry, font_map)
        payload_bytes.append(packed)
        current += len(packed)
    
    # sizes
    table_size = 4 + 4 * n
    payload_len = sum(len(p) for p in payload_bytes)
    block3_size = table_size + payload_len + 4

    # extra add_block bytes
    add_block_bytes = bytes.fromhex(data.get('add_block', ''))
    add_len = len(add_block_bytes)
    
    # Main size: sum of all parts after this field
    # 4 bytes for count1, count1*2 bytes entries,
    # 4 bytes for count2, count2*2 bytes entries,
    # 4 bytes for block3_size, block3_size bytes,
    # plus add_block length
    main_size = (
        4 + count1 * 2 + len(block1_align)
        + 4 + count2 * 2
        + 4 + block3_size
        #+ add_len
    )
    out += struct.pack('<I', main_size)
    
    print("main_size:", main_size)
    print("block_1:", 4 + count1 * 2)
    print("block_2:", 4 + count2 * 2)
    print("block_3:", block3_size)
    
    # Write Block 1
    out += struct.pack('<I', count1)
    for val in entries1[:count1]:
        out += struct.pack('<H', val)
    out += block1_align
    
    # Write Block 2
    out += struct.pack('<I', count2)
    for val in entries2[:count2+1]:
        out += struct.pack('<H', val)
    
    # Write Block 3
    out += struct.pack('<I', block3_size)
    out += struct.pack('<I', table_size)
    for off in offsets:
        out += struct.pack('<I', off)
    for packed in payload_bytes:
        out += packed
        
    print(len(out), len(out))
    out += b'\00' * (len(out) % 2)
    print(len(out))
    
    out += b'\00' * 4
    print(len(out))
    # Add extra block
    out += add_block_bytes
    print(len(out))
    
    return out

def normalize_text(text):
    replacements = {
        '?': '？',
        '!': '！',
        ',': '，',
        '.': '．',
        '"': '”',
        "'": '”',
        ":": '：',
        # ':': '：',
        # ';': '；',
        # '(': '（',
        # ')': '）',
        # '[': '［',
        # ']': '］',
        # '{': '｛',
        # '}': '｝',
        # '<': '＜',
        # '>': '＞',
        # '/': '／',
        # '\\': '＼',
        # '|': '｜',
        # '*': '＊',
        
        # '\'': '＇',
        # '`': '｀',
        # '~': '～',
        # '&': '＆',
        # '#': '＃',
        # '%': '％',
        # '^': '＾',
        # '+': '＋',
        # '=': '＝',
        # '-': '－',
        # '_': '＿',
        # '@': '＠',
        # '$': '＄',
    }

    return ''.join(replacements.get(c, c) for c in text)


from openpyxl import load_workbook

def import_from_excel_unescape(filename):
    wb = load_workbook(filename)
    ws = wb.active

    entries = []
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        cell_value = row[0]
        if cell_value is not None:
            # Unescape \\n -> \n
            unescaped_text = cell_value.replace('\\n', '\n')
            entries.append(unescaped_text)

    return entries
    
   
def main():
    parser = argparse.ArgumentParser(description="Pack JSON dialog into binary format")
    parser.add_argument("infile", help="Input JSON file with dialog structure")
    parser.add_argument("outfile", help="Output binary dialog file")
    parser.add_argument("--excel", help="Path to excel text entries", required=False)
    parser.add_argument(
        "--font_table",
        default="./font/font-table.txt",
        help="Path to font-table.txt"
    )
    parser.add_argument(
        "--ascii_table",
        default="./font/ascii-table.bin",
        help="Path to ascii-table.bin"
    )
    args = parser.parse_args()

    # Load JSON
    with open(args.infile, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Load and insert excel entries
    if args.excel:
        data['block_3']['table_entries'] = import_from_excel_unescape(args.excel)
    
    # Build binary
    bin_data = build_dialog(data, FontMapper(args.ascii_table, args.font_table))

    # Write
    with open(args.outfile, 'wb') as f:
        f.write(bin_data)

    print(f"[+] Binary dialog written to: {args.outfile}")


if __name__ == '__main__':
    main()

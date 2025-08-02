import struct, argparse, json, os

from unpack_spirit import align_4
from font_mapper import FontMapper

import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment
from unpack_spirit import find_signature

#from parse_script import export_to_excel_escape#, parse_script_text

keywords = {}
    
# Script parsers
def parse_database_text(text_bytes, font_map: FontMapper, end_break=False):
    i = 0
    output = []

    def read_next():
        nonlocal i
        if i < len(text_bytes):
            val = text_bytes[i]
            i += 1
            return val
        return None
    
    def is_next():
        nonlocal i
        if i < len(text_bytes):
            return text_bytes[i]
        return None
    
    def read_bracket_content() -> bytes:
        nonlocal i
        end = text_bytes[i:].find(b')')
        text = b''
        if i != -1 and end != -1:
            text = text_bytes[i:i+end+1]
            i += end + 1
        return text

    while i < len(text_bytes):
        byte = read_next()
        if byte is None:
            break
            
        # Control characters
        # 0x00 - [END] End of string/page. Terminates text rendering unless in a keyword block
        # 0x01-0x0b - Two-byte glyphs 
        # 0x0C - Control block
        #        0x0C - [CLUT:X] Set color palette index
        #        0x10 - [IMG:(X,Y,Z)] Insert image to text block
        #        0x14 - [KEYWORD:(X,Y,Z)] Interactive keyword link to text block
        #        0x19 - [END_PAGE] Manually force page break
        #        0x1E - [LINE] Toggle high-bit underline flag
        #        
        # 0x0D - "\n" Force end of current row (line break)
        # 0x20 - " " or [SP:X] if there are more than 2
        #
        # >= 0x20 - One-byte character that are encoded through "ascii-table.bin"
        #           (ASCII-compatible custom mapping)
        
        if byte == 0x00:
            output.append("[END]")
            if end_break:
                break

        elif 0x01 <= byte <= 0x0B and is_next() is not None:
            high = byte
            low = read_next()
            code = (high << 8) | (low if low is not None else 0)
            output.append(font_map.get_char(code))

        elif byte == 0x0C and is_next() is not None:
            subcmd = read_next()
            if subcmd == 0x0C:
                clutIndex = read_next()
                output.append(f"[CLUT:{clutIndex - 0x30}]")
            elif subcmd == 0x10:
                img_params = read_bracket_content()
                output.append(f"[IMG:{img_params.decode('utf-8')}]")
            elif subcmd == 0x14:
                keyw_params = read_bracket_content()
                content = keyw_params.decode('utf-8').strip('()')
                parts = content.split(',')
                if len(parts) == 3:
                    x, y, z = parts
                    keywords[int(x.strip())] = int(z.strip())
                output.append(f"[KEYWORD:{keyw_params.decode('utf-8')}]")
            elif subcmd == 0x19:
                output.append("[END_PAGE]")
            elif subcmd == 0x1e:
                output.append("[LINE]")
            else:
                output.append(f"[UNK3:{subcmd:02X}]")

        elif byte == 0x0D:
            output.append("\n")
        
        elif byte == 0x20:
            count = 1
            while is_next() == 0x20:
                read_next()
                count += 1
            if count == 1:
                char = font_map.get_ascii_char(byte)
                output.append(char)
            else:
                output.append(f"[SP:{count}]")

        # Special Kanji chars
        elif byte == 0xDE:
            output.append(f"゛")
        elif byte == 0xDF:
            output.append(f"゜")

        elif 0x20 <= byte:
            char = font_map.get_ascii_char(byte)
            output.append(char if char else f"[UNK1:{byte:02X}]")
        else:
            output.append(f"[UNK2:{byte:02X}]")
            
    return ''.join(output)
    
def get_text_entry_size(data: bytes) -> int:
    i = 0
    while i < len(data):
        byte = data[i]
        
        # 0x00 — конец строки, если не внутри control-блока
        if byte == 0x00:
            return i + 1
        
        # Двухбайтовые символы: 0x01–0x0B
        if 0x01 <= byte <= 0x0B:
            i += 2
            continue

        # Control block
        elif byte == 0x0C and (i + 1) < len(data):
            subcmd = data[i + 1]
            i += 2
            if subcmd == 0x0C:  # [CLUT:X]
                i += 1
            elif subcmd in (0x10, 0x14):  # [IMG:(...)] or [KEYWORD:(...)]
                end = data[i:].find(b')')
                if end == -1:
                    break  # malformed, bail out
                i += end + 1
            elif subcmd == 0x19:  # [END_PAGE]
                pass  # no additional data
            elif subcmd == 0x1E:  # [LINE]
                pass
            else:
                pass  # unknown, skip only subcmd byte
            continue

        # 0x0D — line break
        elif byte == 0x0D:
            i += 1
            continue

        # 0x20 — space or multiple spaces
        elif byte == 0x20:
            i += 1
            while i < len(data) and data[i] == 0x20:
                i += 1
            continue

        # Special kana (゛゜)
        elif byte in (0xDE, 0xDF):
            i += 1
            continue

        # One-byte characters (ASCII / mapped)
        elif byte >= 0x20:
            i += 1
            continue

        # Unknown byte — treat as one byte
        else:
            i += 1
            continue

    return len(data)


def parse_offsets(data):
    return struct.unpack_from('<IIIII', data, 0)

def parse_entries_table(data, start, end):
    entries = []
    for i in range((end - start) // 22):
        entry_offset = start + i * 22
        entry = {
            "id": struct.unpack_from('<H', data, entry_offset + 0)[0],
            "type": struct.unpack_from('<H', data, entry_offset + 2)[0],
            "reserved": struct.unpack_from('<H', data, entry_offset + 4)[0],
            "topics": struct.unpack_from('<HHHHHHHH', data, entry_offset + 6),
        }
        entries.append(entry)
    return entries

def parse_topics_table(data, offset, next_block_offset, font_map):
    table_size = struct.unpack_from('<I', data, offset)[0] - offset
    table_offsets_count = table_size // 4
    table_offsets = [struct.unpack_from('<I', data, offset + i * 4)[0] for i in range(table_offsets_count)]

    print(offset, next_block_offset)
    print(table_size, table_offsets_count)

    # Вычисляем размеры на основе отсортированных смещений
    sorted_offsets = sorted(table_offsets)
    offset_to_size = {}
    block_size = next_block_offset - offset

    for i, entry_offset in enumerate(sorted_offsets):
        if i + 1 < len(sorted_offsets):
            size = sorted_offsets[i + 1] - entry_offset
        else:
            size = block_size - (entry_offset - offset)
        offset_to_size[entry_offset] = size

    table_entries = []
    raw_table_entries = []

    for idx, entry_offset in enumerate(table_offsets):
        size = offset_to_size[entry_offset]
        entry_data = data[entry_offset:entry_offset + size]
        raw_table_entries.append(entry_data.hex())

        sorted_index = sorted_offsets.index(entry_offset)
        title = entry_data[4:4 + get_text_entry_size(entry_data[4:])]
        #print(title)
        #if sorted_index == 397:
        #    print(entry_data, title, parse_database_text(entry_data[:4], font_map), parse_database_text(title, font_map))
            
        table_entries.append({       
            "sorted_index": sorted_index+1,         
            "search_key" : parse_database_text(entry_data[:4], font_map),
            "title" : parse_database_text(title, font_map),
            "pages" : parse_database_text(entry_data[4 + len(title):], font_map, True)
        })

    return {
        "offsets": table_offsets,
        "entries": table_entries,
        #"raw_entries": raw_table_entries
    }

def parse_keyword_table(data, offset, next_block_offset, font_map):
    table_size = struct.unpack_from('<I', data, offset)[0] - offset
    table_offsets_count = table_size // 4
    table_offsets = [struct.unpack_from('<I', data, offset + i * 4)[0] for i in range(table_offsets_count)]
    block_size = next_block_offset - offset

    print(offset, next_block_offset)
    print(table_size, table_offsets_count)
    print(block_size)

    table_entries = []
    raw_table_entries = []

    for i, entry_offset in enumerate(table_offsets):
        if i + 1 < len(table_offsets):
            size = table_offsets[i + 1] - entry_offset
        else:
            size = block_size - (entry_offset - offset)
        entry_data = data[entry_offset:entry_offset + size]
        raw_table_entries.append(entry_data.hex())
        definition_id = keywords.get(i, "")
        keyword_type = "keyword" if definition_id else "definition"
        table_entries.append({
            "id" : i,
            "type" : keyword_type,
            "definition_id" : definition_id,
            "search_key" : parse_database_text(entry_data[:4], font_map),
            "text" : parse_database_text(entry_data[4:], font_map)
        })

    return {
        "offsets": table_offsets,
        "entries": table_entries,
        #"raw_entries": raw_table_entries
    }

def parse_search_table(data, offset, next_block_offset, font_map):
    table_size = struct.unpack_from('<I', data, offset)[0] - offset
    table_offsets_count = table_size // 4
    table_offsets = [struct.unpack_from('<I', data, offset + i * 4)[0] for i in range(table_offsets_count)]
    block_size = next_block_offset - offset

    print(offset, next_block_offset)
    print(table_size, table_offsets_count)
    print(block_size)

    table_entries = []
    raw_table_entries = []

    for i, entry_offset in enumerate(table_offsets):
        if i + 1 < len(table_offsets):
            size = table_offsets[i + 1] - entry_offset
        else:
            size = block_size - (entry_offset - offset)
        entry_data = data[entry_offset:entry_offset + size]
        raw_table_entries.append(entry_data.hex())

        print("Search", entry_offset, size, len(data[entry_offset:next_block_offset]))

        table_entries.append({
            "key_offset" : entry_data[0],
            "count" : entry_data[1],
            "keys" : parse_database_text(entry_data[2:], font_map)
        })

    return {
        "offsets": table_offsets,
        "entries": table_entries,
        #"raw_entries": raw_table_entries
    }

def parse_database(data, font_map):
    """
    Database file structure:
        1. Offsets:
        [16-bytes : 4-bytes] 
            0x00: [32-bit] Entries Info Table
            0x04: [32-bit] Topics Table
            0x08: [32-bit] Keywords Table
            0x0c: [32-bit] Search Table
            0x10: [32-bit] End Raw Data

        2. Entries Info Table: 
        [1936-bytes : 22-bytes]
            0x00: [16-bit] ID
            0x02: [16-bit] Type ID
            0x04: [16-bit] 0x0000
            0x06-0x16: [16-bit] Reference Entry IDs

        3. Topics Table
            1. Offsets table [32-bit]
            2. Topic entry:
                0x00-0x03: [ASCII] Search Key
                0x04-0x..: [ASCII] Title (until [END])
                0x..: [ASCII] Pages (String separeted by [END_PAGE])

        4. Keywords Table
            1. Offsets table [32-bit]
            2. Keyword entry:
                0x00-0x03: [ASCII] Search Key
                0x04-0x..: [ASCII] Text

        5. Search Table
            1. Offsets table [32-bit]
            2. Search entry:
                0x00: [8-bit] Key Offset
                0x01: [8-bit] Count
                0x02: [ASCII] Keys (String separeted by [END])
    """

    offsets = parse_offsets(data)
    print("Offsets:", offsets)

    entries_table = parse_entries_table(data, offsets[0], offsets[1])
    print(f"Entries info table count: {len(entries_table)}")

    topics_table = parse_topics_table(data, offsets[1] + 4, offsets[2], font_map)
    print(f"Topics table count: {len(topics_table['offsets'])}")

    keywords_table = parse_keyword_table(data, offsets[2], offsets[3], font_map)
    print(f"Keywords table count: {len(keywords_table['offsets'])}")

    search_table = parse_search_table(data, offsets[3], offsets[4], font_map)
    print(f"Search table count: {len(search_table['offsets'])}")

    return {
        "offsets": offsets,
        "entries_table": entries_table,
        "topics_table": topics_table,
        "keywords_table": keywords_table,
        "search_table": search_table,
        "end_raw_data": data[offsets[4]:].hex(),
    }

def export_to_excel_escape(data, filename, default_row_height=15):
    wb = Workbook()
    wb.remove(wb.active)
    ws_topics = wb.create_sheet("topics")

    for idx, entry in enumerate(data["topics_table"]["entries"], start=1):
        ws_topics.cell(row=idx, column=1, value=entry["search_key"])
        ws_topics.cell(row=idx, column=2, value=entry["title"])

        # Escape line breaks \n -> \\n
        escaped_text = entry["pages"].replace('\n', '\\n')
        ws_topics.cell(row=idx, column=3, value=escaped_text)

        # Set the standard row height
        ws_topics.row_dimensions[idx].height = default_row_height

    ws_keywords = wb.create_sheet("keywords")

    for idx, entry in enumerate(data["keywords_table"]["entries"], start=1):
        ws_keywords.cell(row=idx, column=1, value=entry["search_key"])
        escaped_text = entry["text"].replace('\n', '\\n')
        ws_keywords.cell(row=idx, column=2, value=escaped_text)
        ws_keywords.row_dimensions[idx].height = default_row_height

    ws_search = wb.create_sheet("search")

    for idx, entry in enumerate(data["search_table"]["entries"], start=1):
        escaped_text = entry["keys"].replace('\n', '\\n')
        ws_search.cell(row=idx, column=1, value=escaped_text)
        ws_search.row_dimensions[idx].height = default_row_height

    wb.save(filename)

def main():
    global keywords

    parser = argparse.ArgumentParser(description="Parse database file")
    parser.add_argument("file", help="Input .database file")
    parser.add_argument("--json_out", help="Path to save parsed data (.json)")
    parser.add_argument("--excel_out", help="Path to save Excel entries (.xlsx)")
    parser.add_argument("--font_table", default="./font/font-table.txt", help="Path to font-table.txt")
    parser.add_argument("--ascii_table", default="./font/ascii-table.bin", help="Path to ascii-table.bin")
    args = parser.parse_args()

    input_dir = os.path.dirname(args.file)
    base_name = os.path.splitext(os.path.basename(args.file))[0]
    if not args.json_out:
        args.json_out = os.path.join(input_dir, base_name + ".json")
    if not args.excel_out:
        args.excel_out = os.path.join(input_dir, base_name + ".xlsx")

    with open(args.file, 'rb') as f1:
        data = f1.read()

    script_type = find_signature(data)
    if script_type not in ["database"]:
        print("Wrong file type! Not .database")
        return

    parsed_data = parse_database(data, FontMapper(args.ascii_table, args.font_table))

    with open(args.json_out, 'w', encoding='utf-8') as out:
        json.dump(parsed_data, out, indent=2, ensure_ascii=False)
    print("[+] JSON saved to:", args.json_out)

    export_to_excel_escape(parsed_data, args.excel_out)
    print("[+] Excel saved to:", args.excel_out)


if __name__ == '__main__':
    main()